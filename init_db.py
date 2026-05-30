import sqlite3
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

def init_db():
    conn = sqlite3.connect('erp_system.db')
    cursor = conn.cursor()

    # 1. Crear tabla de Usuarios (Login)
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL
    )
    ''')
    
    # Insertar 3 usuarios predefinidos
    users = [
        ('admin', 'admin123', 'Administrador'),
        ('ventas', 'ventas123', 'Vendedor'),
        ('bodega', 'bodega123', 'Bodeguero')
    ]
    cursor.executemany('INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)', users)

    # 2. Crear tabla de Inventario (Maestro de Productos)
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS inventory (
        id_sku TEXT PRIMARY KEY,
        name TEXT,
        variant TEXT,
        type TEXT,
        provider TEXT,
        initial_stock INTEGER DEFAULT 0,
        current_stock INTEGER DEFAULT 0,
        price REAL DEFAULT 0
    )
    ''')

    # 3. Crear tabla de Ventas
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        order_number TEXT,
        payment_method TEXT,
        status TEXT,
        client_name TEXT,
        client_id TEXT,
        client_phone TEXT,
        client_address TEXT,
        client_city TEXT,
        id_sku TEXT,
        quantity INTEGER,
        channel TEXT,
        unit_price REAL,
        total_price REAL,
        FOREIGN KEY (id_sku) REFERENCES inventory(id_sku)
    )
    ''')

    # 4. Crear tabla de CRM Clientes (Vista materializada o tabla)
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        client_id TEXT,
        phone TEXT,
        address TEXT,
        city TEXT,
        total_spent REAL DEFAULT 0,
        total_orders INTEGER DEFAULT 0,
        last_purchase TEXT
    )
    ''')

    # Cargar datos desde Excel
    excel_path = 'upload/BODEGA_SISTEMADEGESTION_2026_MEJORADO.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Cargar Inventario
    ws_inv = wb['📦 INVENTARIO']
    inventory_data = []
    for row in ws_inv.iter_rows(min_row=5, values_only=True):
        r = list(row)
        if r[0] is not None:
            sku = str(r[0])
            name = str(r[1]) if r[1] else ''
            variant = str(r[2]) if r[2] else ''
            type_prod = str(r[5]) if r[5] else ''
            provider = str(r[6]) if r[6] else ''
            
            # Stock inicial y actual (asumimos que la suma de stocks físicos es el inicial/actual para la migración)
            stock = 0
            if r[9] is not None: stock += int(r[9])  # Ini Bogota
            if r[10] is not None: stock += int(r[10]) # Ini Medellin
            
            inventory_data.append((sku, name, variant, type_prod, provider, stock, stock, 0))
            
    cursor.executemany('''
    INSERT OR REPLACE INTO inventory (id_sku, name, variant, type, provider, initial_stock, current_stock, price)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', inventory_data)

    # Cargar Ventas
    ws_v = wb['🛒 VENTAS']
    sales_data = []
    clients_dict = {} # Para construir el CRM dinámicamente
    
    for row in ws_v.iter_rows(min_row=6, values_only=True):
        r = list(row)
        if r[0] is not None and r[3] is not None:
            # Parsear fecha
            date_val = r[0]
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)
                
            order_number = str(r[3])
            payment_method = str(r[4]) if r[4] else ''
            status = str(r[5]) if r[5] else ''
            client_name = str(r[6]) if r[6] else 'Consumidor Final'
            client_id = str(r[7]) if r[7] else ''
            client_phone = str(r[8]) if r[8] else ''
            client_address = str(r[9]) if r[9] else ''
            client_city = str(r[10]) if r[10] else ''
            id_sku = str(r[12]) if r[12] else ''
            
            try:
                quantity = int(r[14]) if r[14] else 0
            except:
                quantity = 0
                
            channel = str(r[15]) if r[15] else ''
            
            try:
                unit_price = float(r[16]) if r[16] else 0.0
            except:
                unit_price = 0.0
                
            try:
                total_price = float(r[20]) if r[20] else (unit_price * quantity)
            except:
                total_price = unit_price * quantity
                
            sales_data.append((date_str, order_number, payment_method, status, client_name, client_id, 
                              client_phone, client_address, client_city, id_sku, quantity, channel, 
                              unit_price, total_price))
                              
            # Construir CRM
            if client_name not in clients_dict:
                clients_dict[client_name] = {
                    'client_id': client_id,
                    'phone': client_phone,
                    'address': client_address,
                    'city': client_city,
                    'total_spent': 0.0,
                    'total_orders': set(),
                    'last_purchase': date_str
                }
            
            clients_dict[client_name]['total_spent'] += total_price
            clients_dict[client_name]['total_orders'].add(order_number)
            if date_str > clients_dict[client_name]['last_purchase']:
                clients_dict[client_name]['last_purchase'] = date_str

    cursor.executemany('''
    INSERT INTO sales (date, order_number, payment_method, status, client_name, client_id, 
                       client_phone, client_address, client_city, id_sku, quantity, channel, 
                       unit_price, total_price)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', sales_data)
    
    # Cargar CRM
    crm_data = []
    for name, data in clients_dict.items():
        crm_data.append((name, data['client_id'], data['phone'], data['address'], data['city'], 
                         data['total_spent'], len(data['total_orders']), data['last_purchase']))
                         
    cursor.executemany('''
    INSERT OR REPLACE INTO clients (name, client_id, phone, address, city, total_spent, total_orders, last_purchase)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', crm_data)

    # Actualizar stock actual basado en ventas (salidas)
    # Solo restamos las ventas completadas
    cursor.execute('''
    UPDATE inventory 
    SET current_stock = initial_stock - (
        SELECT COALESCE(SUM(quantity), 0) 
        FROM sales 
        WHERE sales.id_sku = inventory.id_sku AND status = 'Completado'
    )
    ''')

    conn.commit()
    conn.close()
    print("Base de datos inicializada correctamente.")

if __name__ == '__main__':
    init_db()
